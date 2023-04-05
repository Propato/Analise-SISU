from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
# from pandas.io.formats.style import Styler
import string
    
def cores(df, regiao):
    # Define colors
    red = '#ff2800'
    green = '#5cb800'
    blue = '#39A7FA'
    lilac = '#ccccff'
    style = 'background-color: '
    properties = {
        'border': '1px solid black',
        'text-align': 'center',
        'vertical-align': 'middle'
        }

    df = df.style\
        .apply(lambda row: [(style + lilac) if (row['DS_REGIAO'].lower() == regiao and row['APROVAÇÃO'] == True) else '' for col in row], axis=1)\
        .apply(lambda row: [(style + green) if (row['APROVAÇÃO'] == True) else (style + red) for col in row], axis=1, subset=['SG_IES', 'APROVAÇÃO'])\
        .apply(lambda row: [(style + blue) if (row['DS_REGIAO'].lower() == regiao) else '' for col in row], axis=1, subset=['DS_REGIAO'])\
        .set_properties(**properties)
        
    # df_style = Styler(df, uuid_len=0, cell_ids=False)
    # df_style = df_style\
    #     .apply(lambda row: [(style + lilac) if (row['DS_REGIAO'].lower() == regiao and row['APROVAÇÃO'] == True) else '' for col in row], axis=1)\
    #     .apply(lambda row: [(style + green) if (row['APROVAÇÃO'] == True) else (style + red) for col in row], axis=1, subset=['SG_IES', 'APROVAÇÃO'])\
    #     .apply(lambda row: [(style + blue) if (row['DS_REGIAO'].lower() == regiao) else '' for col in row], axis=1, subset=['DS_REGIAO'])\
    #     .set_properties(**properties)

    print('Cores processadas.')
    return df

def layout2(df, ws, notas):
    df_approval = df.data['APROVAÇÃO']
    total_courses = df_approval.size
    approved = df_approval.loc[df_approval == True].size
    
    ws[1][-1].value = f'APROVAÇÃO ({approved/total_courses:.2%})'
    # print([i.value for i in ws[2]])
    # print(ws.column_dimensions['B'].width)
    print(ws.column_dimensions)
    print()
    for column in ws.iter_cols():
        print(column[0].column)
        print(column[0].coordinate)
        print(type(column[0].coordinate))
        print(column[0].coordinate)
    return

    comentarios = ['Código da Instituição de Ensino Superior', 'Nome da Instituição de Ensino Superior', 'Sigla da Instituição de Ensino Superior', 'Nome do Campus', 'Estado da Instituição de Ensino Superior','Município do Campus',  'Região da Instituição de Ensino Superior', 'Código do curso', 'Nome do curso', 'Grau do curso', 'Turno do curso', 'Periodicidade da abertura de vagas através do SISU', 'Número de vagas totais', 'Número de vagas na cota', 'Número de inscritos totais na cota', 'Tipo de cota', f'Peso da Redação\nNota: {notas[0]}', f'Peso de linguagens\nNota: {notas[1]}', f'Peso de matemática\nNota: {notas[2]}', f'Peso das ciências humanas\nNota: {notas[3]}', f'Peso das ciências da natureza\nNota: {notas[4]}', 'Nota de corte para a chamada regular na cota', 'Suas notas ponderadas', 'Aprovação na vaga, se as suas notas são maiores (VERDADEIRO) ou menores (FALSO) que a nota de corte e a %% de aprovação']

    j=0
    for A in list(string.ascii_uppercase):
    
        if type(ws[f'{A}2'].value) == float or type(ws[f'{A}2'].value) == int:
            ws.column_dimensions[A].width = 12
        else:
            ws.column_dimensions[A].width = 20

        h = len(comentarios[j])*2
        compri = len(comentarios[j])*4

        if h < 50:
            h=50
        if h>110:
            h=110
        if compri < 130:
            compri=130

        ws[f'{A}1'].comment = Comment(comentarios[j], 'David Propato', h, compri)

        j+=1
        if j>=df.shape[1]:
            break
    
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['D'].width = 30

    if not j>=df.shape[1]:
        for A in list(string.ascii_uppercase):
    
            if type(ws[f'A{A}2'].value) == float or type(ws[f'A{A}2'].value) == int:
                ws.column_dimensions[f'A{A}'].width = 12
            else:
                ws.column_dimensions[f'A{A}'].width = 20

            h = len(comentarios[j])*2
            compri = len(comentarios[j])*4

            if h < 50:
                h=50
            if h>110:
                h=110
            if compri < 130:
                compri=130

            ws[f'A{A}1'].comment = Comment(comentarios[j], 'David Propato', h, compri)

            j+=1
            if j>=df.shape[1]:
                break

    print('Configurações de layout processadas.')

def layout(wb, df, dic_df, curso, notas):

    ws0 = wb['Dicionário de dados']
    ws1 = wb[curso]
    
    

    borda = Border(right=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'))
    alinhamento = Alignment(horizontal='center', vertical='center')

    comentarios = ['Código da Instituição de Ensino Superior', 'Nome da Instituição de Ensino Superior', 'Sigla da Instituição de Ensino Superior', 'Nome do Campus', 'Estado da Instituição de Ensino Superior','Município do Campus',  'Região da Instituição de Ensino Superior', 'Código do curso', 'Nome do curso', 'Grau do curso', 'Turno do curso', 'Periodicidade da abertura de vagas através do SISU', 'Número de vagas totais', 'Número de vagas na cota', 'Número de inscritos totais na cota', 'Tipo de cota', f'Peso da Redação\nNota: {notas[0]}', f'Peso de linguagens\nNota: {notas[1]}', f'Peso de matemática\nNota: {notas[2]}', f'Peso das ciências humanas\nNota: {notas[3]}', f'Peso das ciências da natureza\nNota: {notas[4]}', 'Nota de corte para a chamada regular na cota', 'Suas notas ponderadas', 'Aprovação na vaga, se as suas notas são maiores (VERDADEIRO) ou menores (FALSO) que a nota de corte']

    j=0
    for A in list(string.ascii_uppercase):
    
        if type(ws1[f'{A}2'].value) == float or type(ws1[f'{A}2'].value) == int:
            ws1.column_dimensions[A].width = 12
        else:
            ws1.column_dimensions[A].width = 20

        h = len(comentarios[j])*2
        compri = len(comentarios[j])*4

        if h < 50:
            h=50
        if h>110:
            h=110
        if compri < 130:
            compri=130

        ws1[f'{A}1'].comment = Comment(comentarios[j], 'David Propato', h, compri)
    
        for i in range(1, df.shape[0]+2):
            ws1[f'{A}{i}'].border = borda
            ws1[f'{A}{i}'].alignment = alinhamento

        j+=1
        if j>=df.shape[1]:
            break
    
    ws1.column_dimensions['B'].width = 42
    ws1.column_dimensions['D'].width = 30

    if not j>=df.shape[1]:
        for A in list(string.ascii_uppercase):
    
            if type(ws1[f'A{A}2'].value) == float or type(ws1[f'A{A}2'].value) == int:
                ws1.column_dimensions[f'A{A}'].width = 12
            else:
                ws1.column_dimensions[f'A{A}'].width = 20

            h = len(comentarios[j])*2
            compri = len(comentarios[j])*4

            if h < 50:
                h=50
            if h>110:
                h=110
            if compri < 130:
                compri=130

            ws1[f'A{A}1'].comment = Comment(comentarios[j], 'David Propato', h, compri)

            for i in range(1, df.shape[0]+2):
                ws1[f'A{A}{i}'].border = borda
                ws1[f'A{A}{i}'].alignment = alinhamento

            j+=1
            if j>=df.shape[1]:
                break

    alin = Alignment(horizontal='left', vertical='distributed')
    borda = Border(right=Side(border_style='thin', color='00000000'))

    for i in range(1, dic_df.shape[0]+2):
        ws0[f'A{i}'].alignment = alinhamento
        ws0[f'B{i}'].alignment = alin

        if i > 1:
            ws0[f'A{i}'].border = borda
            ws0[f'B{i}'].border = borda

        h = int(len(ws0[f'B{i}'].value)/183)

        if len(ws0[f'B{i}'].value)%183:
            h+=1
        '    '
        ws0.row_dimensions[i].height = 15*h

    ws0[f'A{dic_df.shape[0]+2}'].border = Border(top=Side(border_style='thin', color='00000000'))
    ws0[f'B{dic_df.shape[0]+2}'].border = Border(top=Side(border_style='thin', color='00000000'))

    ws0.column_dimensions['A'].width = 32
    ws0.column_dimensions['B'].width = 165

    print('Configurações de layout processadas.')